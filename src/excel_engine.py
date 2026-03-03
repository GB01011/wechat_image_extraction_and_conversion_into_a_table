"""
Excel 生成和导出模块

功能：
- 将结构化数据写入 Excel 文件
- 支持动态列名识别
- 自动格式化和美化
- 支持自定义文件名和输出路径
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from src.logger import logger


class ExcelGenerator:
    """Excel 文件生成和导出模块"""
    
    # 表头样式配置（固定）
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 数据样式配置（固定）
    DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    DATA_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, output_dir="data/output_excel"):
        """
        初始化 Excel 生成器
        
        Args:
            output_dir: 输出目录
        """
        self.output_dir = output_dir
        
        # 确保输出目录存在
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logger.info(f"✓ 创建输出目录: {output_dir}")
        
        logger.info(f"✓ Excel 生成器初始化完成，输出目录: {output_dir}")
    
    def generate(self, data_list, source_name="WeChat", timestamp=None, columns=None):
        """
        生成 Excel 文件（支持动态列名）
        
        处理步骤：
        1. 验证数据
        2. 确定列名（动态或指定）
        3. 创建 DataFrame
        4. 格式化表格
        5. 保存文件
        
        Args:
            data_list: 数据列表（包含字典）
            source_name: 数据来源名称（用于文件名）
            timestamp: 时间戳（可选，用于文件名）
            columns: 指定的列名列表（如果为 None，则从数据中自动提取）
        
        Returns:
            生成的 Excel 文件路径，失败返回 None
        """
        # 1. 验证数据
        if not data_list:
            logger.error("❌ 数据列表为空，无法生成 Excel")
            return None
        
        if not isinstance(data_list, list):
            logger.error(f"❌ 数据格式错误，期望列表但得到 {type(data_list)}")
            return None
        
        try:
            # 2. 确定列名
            if columns is None:
                # 从第一条数据自动提取列名
                if isinstance(data_list[0], dict):
                    columns = list(data_list[0].keys())
                else:
                    logger.error("❌ 无法从数据中提取列名，请提供 columns 参数")
                    return None
            
            logger.info(f"✓ 列名确定: {columns}")
            
            # 3. 创建 DataFrame
            df = pd.DataFrame(data_list)
            
            # 确保所有指定的列都存在
            for col in columns:
                if col not in df.columns:
                    df[col] = ""
            
            # 只保留指定的列，并按照指定的顺序排列
            df = df[columns]
            
            logger.info(f"✓ 已创建 DataFrame: {len(df)} 行 x {len(df.columns)} 列")
            
            # 4. 生成文件名
            if not timestamp:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 清理文件名中的特殊字符
            source_name_clean = self._clean_filename(source_name)
            file_name = f"表格_{source_name_clean}_{timestamp}.xlsx"
            save_path = os.path.join(self.output_dir, file_name)
            
            # 5. 保存 Excel 文件
            df.to_excel(save_path, index=False, sheet_name="数据")
            logger.info(f"✓ Excel 文件已保存: {save_path}")
            
            # 6. 应用样式格式化
            try:
                self._format_excel(save_path, columns)
                logger.info(f"✓ Excel 格式化完成")
            except Exception as e:
                logger.warning(f"⚠ Excel 格式化失败（不影响数据）: {e}")
            
            # 7. 输出统计信息
            logger.info(f"========== Excel 生成统计 ==========")
            logger.info(f"  记录数: {len(df)}")
            logger.info(f"  字段数: {len(df.columns)}")
            logger.info(f"  文件大小: {os.path.getsize(save_path) / 1024:.1f} KB")
            logger.info(f"  保存路径: {save_path}")
            logger.info(f"==================================")
            
            return save_path
        
        except Exception as e:
            logger.error(f"❌ 生成 Excel 失败: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return None
    
    @staticmethod
    def _format_excel(file_path, columns):
        """
        对 Excel 文件进行格式化处理

        处理内容：
        - 设置列宽（动态，包括空列名处理）
        - 格式化表头
        - 格式化数据行
        - 冻结表头
        
        Args:
            file_path: Excel 文件路径
            columns: 列名列表（用于动态设置列宽）
        """
        from openpyxl import load_workbook
        
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        
        # 1. 动态设置列宽
        for col_num, column_name in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            
            # 处理空列名（使用默认宽度）
            if not column_name or column_name.strip() == "":
                width = 10  # 空列名使用默认宽度
            else:
                # 根据列名长度动态计算列宽，最小 12，最大 30
                width = min(max(len(str(column_name)) + 2, 12), 30)
            
            worksheet.column_dimensions[col_letter].width = width
        
        # 2. 格式化表头（第1行）
        for cell in worksheet[1]:
            cell.fill = ExcelGenerator.HEADER_FILL
            cell.font = ExcelGenerator.HEADER_FONT
            cell.alignment = ExcelGenerator.HEADER_ALIGNMENT
            cell.border = ExcelGenerator.DATA_BORDER
        
        # 3. 格式化数据行
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2), 2):
            for cell in row:
                cell.alignment = ExcelGenerator.DATA_ALIGNMENT
                cell.border = ExcelGenerator.DATA_BORDER
        
        # 4. 冻结表头
        worksheet.freeze_panes = "A2"
        
        # 5. 保存格式化后的文件
        workbook.save(file_path)
    
    @staticmethod
    def _clean_filename(filename):
        """
        清理文件名中的特殊字符
        
        Args:
            filename: 原始文件名
        
        Returns:
            清理后的文件名
        """
        # 移除Windows不允许的字符
        invalid_chars = r'<>:"/\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '_')
        
        # 限制长度
        filename = filename[:50]
        
        return filename
